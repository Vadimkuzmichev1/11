import json
from openpyxl import Workbook
from datetime import datetime
import math

# Create a new Excel workbook
workbook = Workbook()

# Create a new sheet
sheet = workbook.active

# Get the current date and time
current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

# Create the filename with the current date and time
filename = f"stable_{current_datetime}.xlsx"

with open("wallets.txt", "r") as f:
    wallets = [row.strip() for row in f if row.strip()]

unformatted_contract_addresses = [
    "0x55d398326f99059fF775485246999027B3197955",
    "0xc2132D05D31c914a87C6611C10748AEb04B58e8F",
    "0xFd086bC7CD5C481DCC9C85ebE478A1C0b69FCbb9",
    "0x94b008aA00579c1307B0EF2c499aD98a8ce58e58",
    "0x9702230A8Ea53601f5cD2dc00fDBc13d4dF4A8c7",
    "0x049d68029688eAbF473097a2fC38ef61633A3C7A",
    "0x8AC76a51cc950d9822D68b83fE1Ad97B32Cd580d",
    "0x2791Bca1f2de4661ED88A30C99A7a9449Aa84174",
    "0xFF970A61A04b1cA14834A43f5dE4533eBDDB5CC8",
    "0x7F5c764cBc14f9669B88837ca1490cCa17c31607",
    "0xB97EF9Ef8734C71904D8002F8b6Bc66Dd9c48a6E",
    "0x04068DA6C83AFCFA0e13ba15A6696662335D5B75",

]
contract_addresses = []

for address in unformatted_contract_addresses:
    contract_addresses.append(address.lower())


unformated_hot_wallets = [
    "0x161bA15A5f335c9f06BB5BbB0A9cE14076FBb645",
    "0x8894E0a0c962CB723c1976a4421c95949bE2D4E3",
    "0x3c783c21a0383057D128bae431894a5C19F9Cf06",
    "0xdccF3B77dA55107280bd850ea519DF3705D1a75a",
    "0xa180Fe01B906A1bE37BE6c534a3300785b20d947",
    "0x73f5ebe90f27B46ea12e5795d16C4b408B19cc6F",
    "0x29bDfbf7D27462a2d115748ace2bd71A2646946c",
    "0x01C952174C24E1210d26961D456A77A39e1F0BB0",
    "0xBD612a3f30dcA67bF60a39Fd0D35e39B7aB80774",
    "0x1FBe2AcEe135D991592f167Ac371f3DD893A508B",
    "0x515b72Ed8a97F42C568D6A143232775018f133C8",
    "0xEB2d2F1b8c558a40207669291Fda468E50c8A0bB",
    "0xe2fc31F816A9b94326492132018C3aEcC4a93aE1",
    "0x9f8c163cBA728e99993ABe7495F06c0A3c8Ac8b9",
    "0x86d2660297c82aC656715e00c979FB5CA65EEcc5",
    "0xe7804c37c13166fF0b37F5aE0BB07A3aEbb6e245",
    "0xf6436829Cf96EA0f8BC49d300c536FCC4f84C4ED",
    "0xB38e8c17e38363aF6EbdCb3dAE12e0243582891D",
    "0xA16F524a804BEaED0d791De0aa0b5836295A2a84",
    "0x06959153B974D0D5fDfd87D561db6d8d4FA0bb0B",
    "0xdE79CE4f78a20b324d057CDb348B558f0C2CeD85",
    "0x0938C63109801Ee4243a487aB84DFfA2Bba4589e",
    "0x7E4aA755550152a522d9578621EA22eDAb204308",
    "0x62383739d68dd0f844103db8dfb05a7eded5bbe6",
    "0x290275e3db66394c52272398959845170e4dcb88",
    "0x505e71695e9bc45943c58adec1650577bca68fd9",
    "0x7043e4e1c4045424858ecbced80989feafc11b36",
    "0x7598e84b2e114ab62cab288ce5f7d5f6bad35bba"
]

hot_wallets = []

for address in unformated_hot_wallets:
    hot_wallets.append(address.lower())


wallet_temp = {

    'total': 0,
    'total_end': 0,
}

wallets_result = {}
project = []
for wallet in wallets:

    with open(f'data/{wallet}.json', 'r') as file:
        wallet_transaction = json.load(file)

    wallets_result[wallet] = wallet_temp.copy()

    for transaction in wallet_transaction:
        transaction["wallet"] = wallet
        project.append(transaction)

        try:

            if transaction["chain"] in ["arb", 'avax', 'matic', 'ftm', 'op', 'bsc']:
                if transaction["receives"]:
                    if transaction["receives"][0]['token_id'].lower() in contract_addresses:
                        if transaction['other_addr'].lower() in hot_wallets:
                            wallets_result[wallet]['total'] += math.ceil(transaction["receives"][0]["amount"])
                            print("receives", wallet, transaction["receives"][0]["amount"], transaction["id"])

            if transaction["sends"] and not transaction["receives"] and len(transaction["sends"]) == 1:
                if transaction["sends"][0]['token_id'] in contract_addresses:
                    wallets_result[wallet]['total_end'] += int(transaction["sends"][0]["amount"])
                    print("sends", wallet, transaction["sends"][0]["amount"], transaction["id"])

            if transaction["sends"] and len(transaction["sends"]) == 1: # azuro
                if transaction["sends"][0]['token_id'] in contract_addresses and transaction["sends"][0]['to_addr'] == "0x7043e4e1c4045424858ecbced80989feafc11b36":
                    wallets_result[wallet]['total_end'] += int(transaction["sends"][0]["amount"])
                    print("sends", wallet, transaction["sends"][0]["amount"], transaction["id"])


            if transaction['chain'] == "ftm" and transaction["receives"] and transaction['sends']:
                if transaction["sends"][0]['token_id'] in contract_addresses and transaction["receives"][0]['token_id'] == "ftm":
                    if transaction['other_addr'] == '0x2a71693a4d88b4f6ae6697a87b3524c04b92ab38':
                        wallets_result[wallet]['total_end'] += int(transaction["sends"][0]["amount"])
                        print(wallet, transaction["sends"][0]["amount"], transaction["id"])





            # if transaction['tx'] and transaction["sends"] and len(transaction["sends"]) == 1 and transaction["chain"] == 'ftm':  # ftm вывод через нативку
            #     if transaction["sends"][0]['token_id'] in contract_addresses:
            #         if transaction["receives"][0]['token_id'] == "ftm" and transaction['tx']['to_addr'] == '0x2a71693a4d88b4f6ae6697a87b3524c04b92ab38':
            #             wallets_result[wallet]['total_end'] += int(transaction["sends"][0]["amount"])
            #             print(wallet, transaction["sends"][0]["amount"], transaction["id"])



                # BINANCE



        except Exception as e:
            print(e, "ошибка")
            print(transaction, "ошибка")

    # print(wallet)
    # for title in wallets_result[wallet]:
    #     if wallets_result[wallet][title] != 0:
    #         print(title, wallets_result[wallet][title])
    # print()

headers = list(wallet_temp.keys())
headers.insert(0, "Wallet")

# Write headers to the first row of the sheet
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header)

# Write wallet data to the sheet
for row_num, wallet in enumerate(wallets_result.keys(), 1):
    sheet.cell(row=row_num + 1, column=1, value=wallet)  # Write wallet to the first column
    for col_num, key in enumerate(headers[1:], 2):
        sheet.cell(row=row_num + 1, column=col_num, value=wallets_result[wallet][key])  # Write value to corresponding column

# Save the workbook
workbook.save(filename)

addresses = set()
# for transaction in project:
#     if transaction["id"] == "0x11c2f6c80ff29cf9a9e63a4f2a70f2e13569639bb2cd74524c6a8352e8740aa8":
#
#     # if transaction["sends"] and not transaction["receives"] and len(transaction["sends"]) == 1:
#     #     if transaction["sends"][0]['token_id'] in contract_addresses:
#         # if transaction["receives"]:
#             # if transaction["receives"][0]['token_id'] in contract_addresses and transaction['other_addr'] in hot_wallets:
#             # if transaction["receives"][0]['token_id'] == "arb":
#                 # addresses.add(transaction['other_addr'])
#                 print("wallet:", transaction['wallet'])
#                 print("cate_id:", transaction['cate_id'])
#                 print("chain:", transaction['chain'])
#                 print("id:", transaction['id'])
#                 print("other_addr:", transaction['other_addr'])
#                 print(transaction['other_addr'])
#                 print("project_id:", transaction['project_id'])
#                 print("receives:", transaction['receives'])
#                 print("sends:", transaction['sends'])
#                 print("tx:", transaction['tx'])
#                 print()
#
# print(addresses)

# binance 0x161ba15a5f335c9f06bb5bbb0a9ce14076fbb645, 0x9f8c163cba728e99993abe7495f06c0a3c8ac8b9, 0x86d2660297c82ac656715e00c979fb5ca65eecc5
# okex 0x06959153b974d0d5fdfd87d561db6d8d4fa0bb0b, 0x0938c63109801ee4243a487ab84dffa2bba4589e, 0xa16f524a804beaed0d791de0aa0b5836295a2a84
# okex
