import json
from openpyxl import Workbook
from datetime import datetime

# Create a new Excel workbook
workbook = Workbook()

# Create a new sheet
sheet = workbook.active

# Get the current date and time
current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

# Create the filename with the current date and time
filename = f"result\starknet_results_{current_datetime}.xlsx"

contracts = {
    "starknet_off_bridge": "0xae0ee0a63a2ce6baeeffe56e7714fb4efe48d419",
}

with open("wallets.txt", "r") as f:
    wallets = [row.strip() for row in f if row.strip()]

wallet_temp = {
    "starknet_off_bridge": 0,
    "total_trx": 0,
    "total_eth": 0,
    "total_fee": 0,
}

wallets_result = {}
project = []
for wallet in wallets:

    with open(f'data/{wallet}.json', 'r') as file:
        wallet_transaction = json.load(file)

    wallets_result[wallet] = wallet_temp.copy()

    for transaction in wallet_transaction:
        transaction["wallet"] = wallet
        project.append(transaction)

        try:

            if transaction["tx"]["status"] == 1:
                if transaction["other_addr"] == contracts["starknet_off_bridge"] and transaction["chain"] == "eth":
                    wallets_result[wallet]["starknet_off_bridge"] += 1
                    wallets_result[wallet]["total_trx"] += 1
                    wallets_result[wallet]["total_eth"] += transaction["sends"][0]['amount']
                    wallets_result[wallet]["total_fee"] += transaction["tx"]['eth_gas_fee']
            if transaction["tx"]["status"] == 0:
                if transaction["other_addr"] == contracts["starknet_off_bridge"] and transaction["chain"] == "eth":
                    wallets_result[wallet]["total_fee"] += transaction["tx"]['eth_gas_fee']
                    # print(1)

        except Exception as e:
            print(e)
    # #
    # print(wallet)
    # for title in wallets_result[wallet]:
    #     if wallets_result[wallet][title] != 0:
    #         print(title, wallets_result[wallet][title])
    #         print()

headers = list(wallet_temp.keys())
headers.insert(0, "Wallet")

# Write headers to the first row of the sheet
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header)
#
# # Write wallet data to the sheet
for row_num, wallet in enumerate(wallets_result.keys(), 1):
    sheet.cell(row=row_num + 1, column=1, value=wallet)  # Write wallet to the first column
    for col_num, key in enumerate(headers[1:], 2):
        sheet.cell(row=row_num + 1, column=col_num, value=wallets_result[wallet][key])  # Write value to corresponding column

# Save the workbook
workbook.save(filename)


# counter = 0
# #
# for transaction in project:
#     if transaction["cate_id"] != "approve":
#         if transaction["other_addr"] == "0xae0ee0a63a2ce6baeeffe56e7714fb4efe48d419":
#             # if transaction["other_addr"] == "0x7ee459d7fde8b4a3c22b9c8c7aa52abaddd9ffd5":
#             # if transaction["chain"] == "era":
#             # if transaction["receives"]:
#             #     if transaction["receives"][0]["token_id"] == "0x80115c708e12edd42e504c1cd52aea96c547c05c":
#             # counter += 1
#             print("wallet:", transaction['wallet'])
#             print("cate_id:", transaction['cate_id'])
#             print("chain:", transaction['chain'])
#             print("id:", transaction['id'])
#             print("other_addr:", transaction['other_addr'])
#             print("project_id:", transaction['project_id'])
#             print("receives:", transaction['receives'])
#             print("sends:", transaction['sends'])
#             print("tx:", transaction['tx'])
#             print("time_at:", transaction["time_at"])
#
#             print()
#             # print(transaction)
# print(counter)
