import json
from openpyxl import Workbook
from datetime import datetime

# Create a new Excel workbook
workbook = Workbook()

# Create a new sheet
sheet = workbook.active

# Get the current date and time
current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

# Create the filename with the current date and time
filename = f"result\zksync_results_{current_datetime}.xlsx"




wallet_temp = {
    "total_trx": 0,
    "nonce": 0,
    "first_day": "",
    'last_day': "",
    'days': 0,
    'weeks': 0,
    'months': 0,

    "popup_okx": 0,

    "zksync_lite_bridge": 0,
    "zksync_era_bridge": 0,
    "orbiter_bridge_send": 0,
    "orbiter_bridge_recieved": 0,

    'ZKS_domain_reg': 0,
    'ZKS_domain_set': 0,

    "transfer": 0,

    "syncswap_pool": 0,
    "syncswap_swap": 0,
    "1inch_swap": 0,

    "claim_zkape": 0,

    "bungee": 0,

    "swacefi_swap": 0,

    "mute_swap": 0,
    "mute_pool": 0,

    "element_nft_buy": 0,

    "tevaera_mint_nft": 0,
    "tevaera_buy_kp": 0,

}

wallets_result = {}
project = []


def readable_print(tx):
    print("wallet:", tx['wallet'])
    print("cate_id:", tx['cate_id'])
    print("chain:", tx['chain'])
    print("id:", tx['id'])
    print("other_addr:", tx['other_addr'])
    print("project_id:", tx['project_id'])
    print("receives:", tx['receives'])
    print("sends:", tx['sends'])
    print("tx:", tx['tx'])
    print("time_at:", tx["time_at"])

    print()
    print(tx)
    print()


def confirmed(tx):

    if tx['tx']:
        return tx["tx"]["status"] == 1


def chain(tx, chain):
    return tx["chain"] == chain


def receives(tx, token_id=None, from_addr=None):
    if tx['receives']:
        if token_id:
            return tx['receives'][0]['token_id'] == token_id
        if from_addr:
            return tx['receives'][0]['from_addr'] == from_addr
    else:
        return False

def main():
    with open("wallets.txt", "r") as f:
        wallets = [row.strip() for row in f if row.strip()]

    for wallet in wallets:

        with open(f'data/{wallet}.json', 'r') as file:
            wallet_transaction = json.load(file)

        wallets_result[wallet] = wallet_temp.copy()

        timestamps = []

        for tx in wallet_transaction:

            tx["wallet"] = wallet
            project.append(tx)

            try:
                # считаем только удачные транзакции
                if confirmed(tx):

                    # считаем транзакции в сети ETH
                    if chain(tx, "eth"):

                        # официальный мост zksync lite
                        if tx["other_addr"] == "0xabea9132b05a70803a4e85094fd0e1800777fbef":
                            wallets_result[wallet]["zksync_lite_bridge"] += 1

                        # официальный мост zksync era
                        if tx["other_addr"] == "0x32400084c286cf3e17e7b677ea9583e60a000324":
                            wallets_result[wallet]["zksync_era_bridge"] += 1

                    # считаем все транзакции в сети ERA в том числе апрувы
                    if chain(tx, "era"):
                        wallets_result[wallet]["nonce"] += 1
                        timestamps.append(int(tx["time_at"]))
                        readable_print(tx)

                        # исключаем все транзакции апрувов для подсчета целевых транзакций в сети ERA
                        if tx["cate_id"] != "approve":


                            # считаем все транзакции в сети ERA без апрувов
                            wallets_result[wallet]["total_trx"] += 1

                            # SYNCSWAP SWAP SMTH to ETH
                            if tx["other_addr"] == "0x621425a1ef6abe91058e9712575dcc4258f8d091":
                                wallets_result[wallet]["syncswap_swap"] += 1



                            # SYNCSWAP SWAP OTHER
                            if tx["sends"] and tx["other_addr"] == "0x2da10a1e27bf85cedd8ffb1abbe97e53391c0295" and len(tx["sends"]) == 2 and not tx["receives"]:
                                wallets_result[wallet]["syncswap_swap"] += 1
                                readable_print(tx)


                            # SYNCSWAP SWAP ETH to USDC
                            if receives(tx, from_addr="0x621425a1ef6abe91058e9712575dcc4258f8d091"):
                                wallets_result[wallet]["syncswap_swap"] += 1


                            # SYNCSWAP POOL ETH/USDC
                            if receives(tx, token_id="0x80115c708e12edd42e504c1cd52aea96c547c05c"):
                                wallets_result[wallet]["syncswap_pool"] += 1


                            # CLAIM ZKAPE
                            if tx["other_addr"] == "0x9aa48260dc222ca19bdd1e964857f6a2015f4078":
                                wallets_result[wallet]["claim_zkape"] += 1

                            if tx["other_addr"] in ["0x8b791913eb07c32779a16750e3868aa8495f5964", "0xdfaab828f5f515e104baaba4d8d554da9096f0e4", "0x2c0737aaf530714067396131ee9be9cee4cf09a0"]:
                                wallets_result[wallet]["mute_swap"] += 1

                            # MUTE POOL ETH/USDC
                            if receives(tx, token_id="0xdfaab828f5f515e104baaba4d8d554da9096f0e4"):
                                wallets_result[wallet]["mute_pool"] += 1

                            # 1INCH SWAP
                            if tx["other_addr"] == "0x6e2b76966cbd9cf4cc2fa0d76d24d5241e0abc2f":
                                wallets_result[wallet]["1inch_swap"] += 1

                            # SPACEFI SWAP
                            if tx["other_addr"] == "0xbe7d1fd1f6748bbdefc4fbacafbb11c6fc506d1d":
                                wallets_result[wallet]["swacefi_swap"] += 1

                            # REFIL BUNGEE
                            if tx["other_addr"] == "0x7ee459d7fde8b4a3c22b9c8c7aa52abaddd9ffd5":
                                wallets_result[wallet]["bungee"] += 1

                            # Ввод через мост Orbiter
                            if tx["receives"] and tx["other_addr"] in ["0x80c67432656d59144ceff962e8faf8926599bcf8", "0xe4edb277e41dc89ab076a1f049f4a3efa700bce8"]:
                                wallets_result[wallet]["orbiter_bridge_recieved"] += 1

                            # Вывод через мост Orbiter
                            if tx["sends"] and tx["other_addr"] in ["0x80c67432656d59144ceff962e8faf8926599bcf8", "0xee73323912a4e3772b74ed0ca1595a152b0ef282"]:
                                wallets_result[wallet]["orbiter_bridge_send"] += 1

                            # Пополнение с OKX
                            if tx["receives"] and tx["other_addr"] == "0x888270ff52f486729ef865466340d4eac83a31d6":
                                wallets_result[wallet]["popup_okx"] += 1

                            # ZKN REG DOMAIN
                            if tx["other_addr"] == "0xcbe2093030f485adaaf5b61deb4d9ca8adeae509":
                                wallets_result[wallet]["ZKS_domain_reg"] += 1

                            # ZKN SET DOMAIN
                            if tx["other_addr"] == "0xcbe2093030f485adaaf5b61deb4d9ca8adeae509":
                                wallets_result[wallet]["ZKS_domain_set"] += 1

                            # # TRANSFER
                            if tx["cate_id"] == "send" and not tx["receives"] and tx["sends"][0]["token_id"] == "era":
                                wallets_result[wallet]["transfer"] += 1

                            #  ELEMENT
                            if tx["other_addr"] == "0x64848eefbc2921102a153b08fa64536ae1f8e937":
                                # readable_print(tx)
                                wallets_result[wallet]["element_nft_buy"] += 1

                            #  TEVAERA MINT ID

                            if tx["other_addr"] == "0xd29aa7bdd3cbb32557973dad995a3219d307721f":
                                wallets_result[wallet]["tevaera_mint_nft"] += 1

                            #  TEVAERA MINT NFT
                            if tx["other_addr"] == "0x50b2b7092bcc15fbb8ac74fe9796cf24602897ad":
                                wallets_result[wallet]["tevaera_mint_nft"] += 1


                            #  TEVAERA BUY KP
                            if tx["other_addr"] == "0x9fc20170d613766831f164f1831f4607ae54ff2d":
                                wallets_result[wallet]["tevaera_buy_kp"] += 1




            except Exception as e:
                print(wallet, readable_print(tx), e,)

        if timestamps:
            wallets_result[wallet]["last_day"] = datetime.fromtimestamp(max(timestamps)).strftime("%d.%m.%Y")
            wallets_result[wallet]["first_day"] = datetime.fromtimestamp(min(timestamps)).strftime("%d.%m.%Y")

            # print(timestamps)
            wallets_result[wallet]["days"] = len(set([datetime.fromtimestamp(ts).date() for ts in timestamps]))
            wallets_result[wallet]["weeks"] = len(set([datetime.fromtimestamp(ts).strftime('%Y-%W') for ts in timestamps]))
            wallets_result[wallet]["months"] = len(set([datetime.fromtimestamp(ts).strftime('%Y-%m') for ts in timestamps]))
        # #
        # print(wallet)
        # for title in wallets_result[wallet]:
        #     if wallets_result[wallet][title] != 0:
        #         print(title, wallets_result[wallet][title])
        # print()

    headers = list(wallet_temp.keys())
    headers.insert(0, "Wallet")

    # Write headers to the first row of the sheet
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)
    #
    # # Write wallet data to the sheet
    for row_num, wallet in enumerate(wallets_result.keys(), 1):
        sheet.cell(row=row_num + 1, column=1, value=wallet)  # Write wallet to the first column
        for col_num, key in enumerate(headers[1:], 2):
            sheet.cell(row=row_num + 1, column=col_num, value=wallets_result[wallet][key])  # Write value to corresponding column

    # Save the workbook
    # workbook.save(filename)


main()